"""
Views xuất báo cáo XLSX (streaming)
"""
from datetime import datetime, date, timedelta
from io import BytesIO

from django.utils.encoding import iri_to_uri
from django.http import StreamingHttpResponse
from django.utils import timezone

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework import status
from rest_framework.response import Response

from openpyxl import Workbook

from api.orders.services import OrderService


def _build_summary_sheet(wb, from_date_obj, to_date_obj):
    ws = wb.active if wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active.title == "Sheet" else wb.create_sheet()
    ws.title = "Summary"
    stats = OrderService.get_revenue_statistics(from_date_obj, to_date_obj)
    summary = stats.get('summary', {})
    ws.append(["From", "To", "Total Orders", "Total Revenue", "Total Discount", "Average Order Value"])
    ws.append([
        from_date_obj.isoformat() if from_date_obj else "",
        to_date_obj.isoformat() if to_date_obj else "",
        summary.get('total_orders', 0),
        summary.get('total_revenue', 0),
        summary.get('total_discount', 0),
        summary.get('average_order_value', 0),
    ])


def _build_trend_sheet(wb, from_date_obj, to_date_obj, group_by: str):
    ws = wb.create_sheet(title="Trend")
    data = OrderService.get_revenue_trend(from_date_obj, to_date_obj, group_by)
    labels = data.get('labels', [])
    datasets = data.get('datasets') or []
    series = datasets[0] if datasets else {"data": []}
    values = series.get('data', [])
    ws.append(["Period", "Revenue"])
    for label, value in zip(labels, values):
        ws.append([label, value])


def _build_by_category_sheet(wb, from_date_obj, to_date_obj):
    ws = wb.create_sheet(title="ByCategory")
    data = OrderService.get_revenue_by_category(from_date_obj, to_date_obj)
    labels = data.get('labels', [])
    values = (data.get('datasets') or [{}])[0].get('data', [])
    ws.append(["Category", "Revenue"])
    for label, value in zip(labels, values):
        ws.append([label, value])


def _build_best_selling_sheet(wb, from_date_obj, to_date_obj, limit: int):
    ws = wb.create_sheet(title="BestSelling")
    items = OrderService.get_best_selling_products(from_date_obj, to_date_obj, limit)
    ws.append(["ProductID", "ProductName", "TotalQuantity", "TotalRevenue", "OrderCount"])
    for item in items:
        ws.append([
            item.get('ProductID'),
            item.get('ProductID__ProductName'),
            item.get('total_quantity'),
            item.get('total_revenue'),
            item.get('order_count'),
        ])


def _workbook_to_chunks(wb, chunk_size: int = 4096):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    while True:
        data = buffer.read(chunk_size)
        if not data:
            break
        yield data


@api_view(['GET'])
@permission_classes([AllowAny])
def export_report_xlsx(request):
    """
    GET /api/reports/export.xlsx?from_date&to_date&type=revenue
    Trả về streaming XLSX: Content-Type + Content-Disposition
    Tên file: bao_cao_{ngày xuất báo cáo}.xlsx
    """
    report_type = (request.query_params.get('type') or 'revenue').lower()
    group_by = (request.query_params.get('group_by') or 'day').lower()
    period = (request.query_params.get('period') or '').lower()
    limit_param = request.query_params.get('limit')
    try:
        best_limit = int(limit_param) if limit_param is not None else 10
    except ValueError:
        best_limit = 10

    from_date = request.query_params.get('from_date')
    to_date = request.query_params.get('to_date')

    # Chuẩn hóa phạm vi ngày: ưu tiên from_date/to_date, nếu không có thì dùng period
    def _parse_date(s: str):
        return datetime.strptime(s, '%Y-%m-%d').date()

    from_date_obj = None
    to_date_obj = None
    if from_date or to_date:
        if from_date:
            try:
                from_date_obj = _parse_date(from_date)
            except ValueError:
                return Response({'error': 'from_date phải có định dạng YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
        if to_date:
            try:
                to_date_obj = _parse_date(to_date)
            except ValueError:
                return Response({'error': 'to_date phải có định dạng YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        # Map period -> from_date/to_date theo timezone hiện tại
        today = timezone.localdate()
        if period in ('today', ''):
            from_date_obj = today
            to_date_obj = today
        elif period == 'week':
            start = today - timedelta(days=today.weekday())
            from_date_obj = start
            to_date_obj = today
        elif period == 'month':
            start = today.replace(day=1)
            from_date_obj = start
            to_date_obj = today
        elif period == 'year':
            start = date(today.year, 1, 1)
            from_date_obj = start
            to_date_obj = today
        else:
            # period không hợp lệ
            return Response({'error': 'period không hợp lệ. Hỗ trợ: today|week|month|year hoặc dùng from_date/to_date'}, status=status.HTTP_400_BAD_REQUEST)

    wb = Workbook()

    # Xây sheet theo type
    if report_type == 'revenue':
        _build_summary_sheet(wb, from_date_obj, to_date_obj)
    elif report_type == 'trend':
        _build_trend_sheet(wb, from_date_obj, to_date_obj, group_by)
    elif report_type == 'by_category':
        _build_by_category_sheet(wb, from_date_obj, to_date_obj)
    elif report_type == 'best_selling':
        _build_best_selling_sheet(wb, from_date_obj, to_date_obj, best_limit)
    elif report_type == 'full':
        _build_summary_sheet(wb, from_date_obj, to_date_obj)
        _build_trend_sheet(wb, from_date_obj, to_date_obj, group_by)
        _build_by_category_sheet(wb, from_date_obj, to_date_obj)
        _build_best_selling_sheet(wb, from_date_obj, to_date_obj, best_limit)
    else:
        ws = wb.active
        ws.title = "Report"
        ws.append(["Message"]) 
        ws.append([f"Unsupported report type: {report_type}"])

    # Đặt tên file: bao_cao_{period|from_to}.xlsx
    if request.query_params.get('from_date') or request.query_params.get('to_date'):
        start_str = from_date_obj.isoformat() if from_date_obj else ''
        end_str = to_date_obj.isoformat() if to_date_obj else ''
        suffix = f"{start_str}__{end_str}".strip('_')
    else:
        suffix = (period or 'today')

    filename = f"bao_cao_{suffix}.xlsx"

    response = StreamingHttpResponse(
        _workbook_to_chunks(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{iri_to_uri(filename)}"'
    return response


